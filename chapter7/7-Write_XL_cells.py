import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
#sheet.title = 'Sample Sheet'
wb.create_sheet(index=0, title='First Sheet')
wb.create_sheet(index=1, title='Second Sheet')


sheet = wb["First Sheet"]
sheet['A1'] = 'Hello world!'
sheet['B1'] = 'Column'



wb.save('example_new1.xlsx')